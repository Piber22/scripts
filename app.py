# app.py
import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
import os
import shutil
from datetime import datetime
import tempfile

# -------------------------------------------------
# Configura√ß√£o da p√°gina
# -------------------------------------------------
st.set_page_config(page_title="Gerenciador de Longas", layout="centered")
st.title("üõèÔ∏è Gerenciador de Longas")
st.markdown(
    "Fa√ßa upload do **PDF atual** para atualizar a planilha **Longas.xlsx**.\n"
)

# -------------------------------------------------
# Uploads
# -------------------------------------------------
col1, col2 = st.columns(2)
with col1:
    pdf_novo_file = st.file_uploader("üìÑ PDF Atual (data.pdf)", type="pdf", key="novo")
with col2:
    pdf_antigo_file = st.file_uploader("üìÑ PDF Anterior (opcional)", type="pdf", key="antigo")

# -------------------------------------------------
# Bot√£o de processamento
# -------------------------------------------------
if st.button("üöÄ Atualizar Longas", type="primary"):
    if not pdf_novo_file:
        st.error("Fa√ßa upload do PDF atual.")
    else:
        with st.spinner("Processando..."):
            try:
                # -------------------------------------------------
                # Diret√≥rio tempor√°rio
                # -------------------------------------------------
                temp_dir = tempfile.mkdtemp()
                excel_path   = os.path.join(temp_dir, "Longas.xlsx")
                pdf_path_new = os.path.join(temp_dir, "data.pdf")
                pdf_path_old = os.path.join(temp_dir, "data_anterior.pdf")

                # Salvar PDFs enviados
                with open(pdf_path_new, "wb") as f:
                    f.write(pdf_novo_file.getbuffer())
                if pdf_antigo_file:
                    with open(pdf_path_old, "wb") as f:
                        f.write(pdf_antigo_file.getbuffer())

                # -------------------------------------------------
                # Copiar planilha original do reposit√≥rio
                # -------------------------------------------------
                original_excel = "Longas.xlsx"          # deve estar na raiz do repo
                if not os.path.exists(original_excel):
                    st.error("Planilha original 'Longas.xlsx' n√£o encontrada no reposit√≥rio.")
                    st.stop()
                shutil.copy2(original_excel, excel_path)

                # -------------------------------------------------
                # Fun√ß√µes auxiliares
                # -------------------------------------------------
                def extrair_data_pdf(caminho):
                    """Extrai a primeira data DD/MM/AAAA encontrada na primeira p√°gina."""
                    try:
                        with pdfplumber.open(caminho) as pdf:
                            texto = pdf.pages[0].extract_text()
                            import re
                            m = re.search(r"(\d{2}/\d{2}/\d{4})", texto)
                            return m.group(1) if m else datetime.now().strftime("%d/%m/%Y")
                    except:
                        return datetime.now().strftime("%d/%m/%Y")

                def extrair_pdf(caminho):
                    """Extrai tabelas do PDF e devolve DataFrame com colunas requeridas."""
                    cols = ["Leito", "Atendimento", "Paciente", "Dias de Ocupa√ß√£o"]
                    dfs = []
                    with pdfplumber.open(caminho) as pdf:
                        for page in pdf.pages:
                            tables = page.extract_tables(table_settings={
                                "vertical_strategy": "lines",
                                "horizontal_strategy": "lines",
                                "snap_x_tolerance": 3,
                                "join_x_tolerance": 3,
                                "join_y_tolerance": 3,
                            })
                            for tab in tables:
                                if not tab or not tab[0]:
                                    continue
                                header = [c.strip() if c else "" for c in tab[0]]
                                if not set(cols).issubset(header):
                                    continue
                                df = pd.DataFrame(tab[1:], columns=header)
                                if "M√©trica" in df.columns:
                                    df.drop("M√©trica", axis=1, inplace=True)
                                dfs.append(df[cols])
                    if not dfs:
                        return pd.DataFrame(columns=cols)
                    df_pdf = pd.concat(dfs, ignore_index=True)
                    df_pdf["Dias de Ocupa√ß√£o"] = (
                        pd.to_numeric(df_pdf["Dias de Ocupa√ß√£o"], errors="coerce")
                        .fillna(0)
                        .astype(int)
                    )
                    df_pdf = df_pdf.drop_duplicates("Leito")
                    return df_pdf

                def atualizar_dados(df_novo, data_pdf):
                    """Atualiza aba 'Dados' a partir da linha 6, mant√©m E/F e formata."""
                    wb = load_workbook(excel_path)
                    ws = wb["Dados"]

                    # ---- Data de gera√ß√£o na F2 (DD/MM/AAAA) ----
                    ws["F2"] = datetime.now().strftime("%d/%m/%Y")

                    # ---- Ler observa√ß√µes atuais (E e F) ----
                    obs_dict = {}
                    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
                        if row[0]:
                            e_val = row[4] if len(row) > 4 else ""
                            f_val = row[5] if len(row) > 5 else ""
                            obs_dict[row[0]] = (e_val, f_val)

                    # ---- Limpar linhas a partir da 6 ----
                    if ws.max_row >= 6:
                        ws.delete_rows(6, ws.max_row - 5)

                    # ---- Ordenar: BOX primeiro ----
                    df_novo["BOX"] = df_novo["Leito"].str.startswith("BOX")
                    df_novo = df_novo.sort_values(
                        by=["BOX", "Leito"], ascending=[False, True]
                    ).drop(columns=["BOX"])

                    # ---- Estilos ----
                    borda = Border(
                        left=Side("thin"),
                        right=Side("thin"),
                        top=Side("thin"),
                        bottom=Side("thin"),
                    )
                    centro = Alignment(horizontal="center", vertical="center")
                    esquerda = Alignment(horizontal="left", vertical="center")

                    # ---- Inserir dados ----
                    for r_idx, row in enumerate(df_novo.itertuples(index=False), start=6):
                        leito = row[0]
                        for c_idx, val in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=val)
                            cell.border = borda
                            cell.alignment = centro if c_idx in [1, 2, 4] else esquerda

                        # Restaurar E e F
                        e, f = obs_dict.get(leito, ("", ""))
                        ws.cell(r_idx, 5, e).border = borda
                        ws.cell(r_idx, 5, e).alignment = centro
                        ws.cell(r_idx, 6, f).border = borda
                        ws.cell(r_idx, 6, f).alignment = esquerda

                    # Data do PDF na F3
                    ws["F3"] = data_pdf
                    wb.save(excel_path)

                def atualizar_historico(df_old, df_new, data_pdf, obs_dict):
                    """Adiciona leitos que sa√≠ram do PDF atual na aba Historico."""
                    if not os.path.exists(pdf_path_old):
                        return 0
                    wb = load_workbook(excel_path)
                    ws = wb["Historico"]
                    removidos = set(df_old["Leito"]) - set(df_new["Leito"])
                    if not removidos:
                        return 0

                    df_baixa = df_old[df_old["Leito"].isin(removidos)].copy()
                    df_baixa["Situa√ß√£o"] = df_baixa["Leito"].map(
                        lambda x: obs_dict.get(x, ("", ""))[0]
                    )
                    df_baixa["OBS"] = df_baixa["Leito"].map(
                        lambda x: obs_dict.get(x, ("", ""))[1]
                    )
                    df_baixa["Data de Baixa"] = data_pdf

                    start_row = ws.max_row + 1
                    for r_idx, row in enumerate(df_baixa.itertuples(index=False), start_row):
                        for c_idx, val in enumerate(row, 1):
                            ws.cell(r_idx, c_idx, val)
                    wb.save(excel_path)
                    return len(df_baixa)

                # -------------------------------------------------
                # Execu√ß√£o principal
                # -------------------------------------------------
                data_pdf = extrair_data_pdf(pdf_path_new)
                df_new   = extrair_pdf(pdf_path_new)

                if df_new.empty:
                    st.warning("Nenhuma tabela v√°lida encontrada no PDF.")
                else:
                    # ---- Observa√ß√µes atuais (antes de sobrescrever) ----
                    wb_temp = load_workbook(excel_path)
                    obs_dict = {}
                    for row in wb_temp["Dados"].iter_rows(min_row=6, values_only=True):
                        if row[0]:
                            obs_dict[row[0]] = (
                                row[4] if len(row) > 4 else "",
                                row[5] if len(row) > 5 else "",
                            )

                    atualizar_dados(df_new, data_pdf)

                    baixas = 0
                    if os.path.exists(pdf_path_old):
                        df_old = extrair_pdf(pdf_path_old)
                        baixas = atualizar_historico(df_old, df_new, data_pdf, obs_dict)
                        # Atualiza backup do PDF anterior
                        shutil.copy2(pdf_path_new, pdf_path_old)
                    else:
                        st.info("Primeira execu√ß√£o ‚Äì hist√≥rico ser√° criado na pr√≥xima.")

                    # -------------------------------------------------
                    # Download da planilha gerada
                    # -------------------------------------------------
                    with open(excel_path, "rb") as f:
                        st.download_button(
                            label=f"üì• Baixar Longas.xlsx ({len(df_new)} leitos, {baixas} baixas)",
                            data=f,
                            file_name=f"Longas_{data_pdf.replace('/', '-')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    st.success("Planilha atualizada com sucesso!")

            except Exception as e:
                st.error(f"Erro inesperado: {e}")
